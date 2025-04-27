import gspread
from google.oauth2.service_account import Credentials
from openpyxl.reader.excel import load_workbook
from datetime import datetime
from openpyxl.styles import numbers
import json

def google_sheets_to_excel():

    with open("config.json") as f:
        config = json.load(f)

    sheet_id = config["sheet_id"]
    credentials_file = config["credentials_file"]

    scopes = [
        "https://www.googleapis.com/auth/spreadsheets"
    ]
    try:
        creds = Credentials.from_service_account_file(credentials_file, scopes=scopes)
        client = gspread.authorize(creds)
    except Exception as e:
        print(f'Error with Google Sheets authentication: {e}')
        return

    sheet_id = sheet_id

    try:
        workbook = client.open_by_key(sheet_id)
        worksheet = workbook.worksheet('Purchase Information')
        values = worksheet.get_all_records()
    except Exception as e:
        print(f'Error accessing Google Sheets data: {e}')
        return

    try:
        excel_file = load_workbook("Sales Information.xlsx")
        excel_sheet = excel_file['Sales Information']
    except Exception as e:
        print(f'Error opening Excel file: {e}')
        return

    for index in range(len(values)):
        if values[index]['Processed'] != 'TRUE':

            try:
                customer_name = values[index]["Customer Name"]
                company_name = values[index]["Company Name"]
                customer_email = values[index]["Email"]
                street_address = values[index]["Street Address"]
                city = values[index]["City"]
                product_list = [product.strip() for product in values[index]["Product(s)"].split(',')]
                unit_price_list = [str(unit_price).strip() for unit_price in str(values[index]["Unit Price(s)"]).split(",")]
                quantity_list = [str(quantity).strip() for quantity in str(values[index]["Quantity"]).split(",")]
                credit_terms = values[index]["Credit Term (days)"]
                current_date = datetime.now()
            except KeyError as e:
                print(f'Missing expected column in Google Sheets: {e}')
                continue
            except Exception as e:
                print(f'Error processing product data: {e}')
                continue

            try:
                last_customer_id = excel_sheet[f'M{excel_sheet.max_row}'].value
                customer_id = int(last_customer_id) + 1 if last_customer_id and str(last_customer_id).isdigit() else 1
            except Exception as e:
                print(f'Error processing Customer ID: {e}')
                customer_id = 1

            for new_index in range(len(product_list)):

                try:
                    subtotal = float(unit_price_list[new_index]) * int(quantity_list[new_index])
                    sales_tax = 9 / 100 * subtotal
                    grand_total = subtotal + sales_tax
                    is_processed = "FALSE"
                except ValueError as e:
                    print(f'Error calculating totals: {e}')
                    return

                excel_sheet.append([customer_name, company_name, customer_email, street_address, city,
                                    product_list[new_index], round(float(unit_price_list[new_index]), 2), int(quantity_list[new_index]),
                                    round(subtotal, 2), round(sales_tax, 2), round(grand_total, 2),
                                    current_date, customer_id, credit_terms, is_processed])

                excel_sheet[f'G{excel_sheet.max_row}'].number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
                excel_sheet[f'I{excel_sheet.max_row}'].number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
                excel_sheet[f'J{excel_sheet.max_row}'].number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
                excel_sheet[f'K{excel_sheet.max_row}'].number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
                excel_sheet[f'L{excel_sheet.max_row}'].number_format = numbers.FORMAT_DATE_DDMMYY

            try:
                worksheet.update_cell(index + 2, 11, 'TRUE')
            except Exception as e:
                print(f'Error updating Google Sheets: {e}')
    try:
        excel_file.save("Sales Information.xlsx")
        excel_file.close()
    except Exception as e:
        print(f'Error saving Excel file: {e}')

google_sheets_to_excel()